import os
from openpyxl import Workbook, load_workbook


def save_to_excel(filename, data):
    if os.path.exists(filename): #decyduje, czy dzialac na istniejacym pliku
        wb = load_workbook(filename)
        ws = wb.active


    else:   #decyduje, czy dzialac juz na istniejacym pliku
        wb = Workbook()
        ws = wb.active
        ws.append(list(data.keys())) #["Miejsce","Temp.","Temp. odczuwalna", "Wilgotność","Prędkość wiatru"])

    ws.append(list(data.values()))  #metoda .append powoduje dodanie danych na dole excela (w pierwszej wolnej kolumnie)
    wb.save(filename)



# def create_excel():
#
#     values = [23,21,22,21,19,24,40]
#
#     wb = Workbook()
#     sheet = wb.active
#     sheet.title = "Temperatury"
#
#     sheet.append(["Pomiar nr.","Temperatura", "test"])
#
#     for index, value in enumerate (values,start=1):
#         sheet.append([index, value, f"test{index}"])
#
#
#     wb.save("testowy.xlsx")
#
# create_excel()

